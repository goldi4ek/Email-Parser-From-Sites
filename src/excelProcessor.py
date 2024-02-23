import openpyxl


class ExcelProcessor:
    """
    A class for processing Excel files and extracting emails from websites.

    Args:
        filename (str): The path to the Excel file.
        scraper (Scraper): An instance of the Scraper class used for website scraping.

    Attributes:
        wb (Workbook): The openpyxl Workbook object representing the Excel file.
        scraper (Scraper): An instance of the Scraper class used for website scraping.
        filename (str): The path to the Excel file.

    Methods:
        process_links_and_get_emails: Process the links in the Excel file and extract emails from websites.
    """

    def __init__(self, filename, scraper):
        self.wb = openpyxl.load_workbook(filename)
        self.scraper = scraper
        self.filename = filename

    def process_links_and_get_emails(self):
        """
        Process the links in the Excel file and extract emails from websites.

        This method iterates through each sheet in the Excel file, checks if the "phone" column exists,
        inserts a new "email" column if necessary, and extracts emails from websites listed in the "phone" column.

        Returns:
            None
        """
        for sheet in self.wb.worksheets:
            if sheet.cell(row=1, column=5).value == "phone":
                sheet.insert_cols(idx=5)
                sheet.cell(row=1, column=5).value = "email"
            for row in sheet.iter_rows():
                cell = row[3]
                if cell.value is not None and cell.row != 1 and row[4].value is None:
                    websites_to_check = []
                    websites_in_cell = str(cell.value)
                    websites_in_cell = websites_in_cell.split(",")
                    for website_in_cell in websites_in_cell:
                        websites_to_check.append(website_in_cell.strip())
                    for website_to_check in websites_to_check:
                        emails = []
                        email = self.scraper.check_page(website_to_check.strip())
                        if email is not None:
                            emails.append(email)
                            if len(emails) > 1:
                                emails = ", ".join(emails)
                            else:
                                emails = emails[0]

                            sheet.cell(row=cell.row, column=5).value = emails
                            self.wb.save(self.filename)
