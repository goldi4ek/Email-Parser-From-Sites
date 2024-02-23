from selenium.webdriver import Chrome
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.proxy import ProxyType
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
from bs4 import BeautifulSoup
import requests
import openpyxl
import re


class Scraper:
    def __init__(self):
        # Set up the Chrome driver
        options = Options()
        options.add_experimental_option("excludeSwitches", ["enable-logging"])
        options.add_argument("--headless")  # Run Chrome in headless mode
        service = Service("chromedriver.exe")
        self.driver = Chrome(service=service, options=options)

    # Check the page for an email address
    def check_page(self, url):
        print(f"\nChecking page: {url}:")
        try:
            response = requests.get(f"https://{url}", timeout=5)
            if response.status_code == 200:
                page = BeautifulSoup(response.content, "html.parser")
                email = self.find_email(page)
                if email:
                    print(email)
                    return email
                else:
                    facebook_link = self.find_facebook_link(page)
                    if facebook_link:
                        if self.check_facebook_link(facebook_link):
                            return self.process_facebook_page(facebook_link)
                        else:
                            try:
                                self.create_and_check_fb_link(url)
                            except:
                                print("Wrong facebook link format")
                                return None
                    else:
                        print("No email and facebook found")
                        return None
            else:
                self.create_and_check_fb_link(url)
                print(
                    f"Failed to load page: {url} with status code {response.status_code}"
                )
        except:
            try:
                self.create_and_check_fb_link(url)

            except requests.exceptions.Timeout:
                print(f"Connection timed out for page: {url}")
                return None
            except requests.exceptions.SSLError:
                print(f"SSL error for page: {url}")
                return None
            except Exception as e:
                with open("log.txt", "a") as f:
                    f.write(f"Could not connect to {url}: {e}\n")
                print(f"Some error on page: {url}")
                return None

    def check_facebook_link(self, url):
        facebook_link = url.split("/")
        return len(facebook_link) == 4

    def create_and_check_fb_link(self, url):
        fb_link = self.create_fb_link(url)
        self.process_facebook_page(fb_link)

    def create_fb_link(self, url):
        name_of_fb_page = url.split(".")[0]
        return f"https://www.facebook.com/{name_of_fb_page}"

    # Find the email address on the page
    def find_email(self, page):
        email = page.find("a", href=re.compile(r"mailto:"))
        if email:
            return email["href"].split(":")[1]
        return None

    # Find the Facebook link on the page
    def find_facebook_link(self, page):
        facebook_link = page.find("a", href=re.compile(r"facebook\.com"))
        if facebook_link:
            return facebook_link["href"]
        return None

    # Process the Facebook page
    def process_facebook_page(self, facebook_link):
        self.driver.get(facebook_link)
        WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located((By.TAG_NAME, "body"))
        )

        # Close the modal window of logging in to Facebook
        try:
            close_button = self.driver.find_element(
                By.XPATH,
                "/html/body/div[1]/div/div[1]/div/div[5]/div/div/div[1]/div/div[2]/div/div/div/div[1]",
            )
            close_button.click()
        except:
            pass

        # Get the page source
        page_source = self.driver.page_source

        # Parse the page source with BeautifulSoup
        soup = BeautifulSoup(page_source, "html.parser")

        # Find the block with the contact information
        info_block = soup.find(class_="x78zum5 x1n2onr6 xh8yej3")
        if info_block:
            # Iterate over each div element within the block
            for div in info_block.find_all("div"):
                # Check if the div element contains an email address
                if re.match(
                    r"\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b",
                    div.text,
                ):
                    print(div.text)
                    return div.text
            print("No email found on Facebook page")
            return None
        else:
            print("No contact information found on Facebook page")
            return None

    def close(self):
        self.driver.quit()


class ExcelProcessor:
    def __init__(self, filename, scraper):
        self.wb = openpyxl.load_workbook(filename)
        self.scraper = scraper
        self.filename = filename

    def process_links_and_get_emails(self):
        for sheet in self.wb.worksheets:
            if sheet.cell(row=1, column=5).value == "phone":
                sheet.insert_cols(idx=5)
                sheet.cell(row=1, column=5).value = "email"
            for row in sheet.iter_rows():
                cell = row[3]
                if cell.value is not None and cell.row != 1 and row[4].value is None:
                    websites_to_check = []
                    websites_in_cell = str(cell.value)
                    websites_in_cell = websites_in_cell.split(",")
                    for website_in_cell in websites_in_cell:
                        websites_to_check.append(website_in_cell.strip())
                    for website_to_check in websites_to_check:
                        emails = []
                        email = self.scraper.check_page(website_to_check.strip())
                        if email is not None:
                            emails.append(email)
                            if len(emails) > 1:
                                emails = ", ".join(emails)
                            else:
                                emails = emails[0]

                            sheet.cell(row=cell.row, column=5).value = emails
                            self.wb.save(self.filename)


def main():
    scraper = Scraper()

    excel_processor = ExcelProcessor("USA Services.xlsx", scraper)

    excel_processor.process_links_and_get_emails()

    scraper.close()


if __name__ == "__main__":
    main()